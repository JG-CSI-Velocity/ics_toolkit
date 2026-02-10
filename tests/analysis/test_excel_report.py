"""Tests for exports/excel.py -- Excel report generation."""

import pandas as pd
import pytest

from ics_toolkit.analysis.analyses.base import AnalysisResult
from ics_toolkit.analysis.exports.excel import write_excel_report


@pytest.fixture
def mock_analyses():
    """Create simple mock analysis results."""
    return [
        AnalysisResult(
            name="Test Analysis 1",
            title="Test Title 1",
            df=pd.DataFrame(
                {
                    "Category": ["A", "B", "Total"],
                    "Count": [10, 20, 30],
                    "% of Count": [0.333, 0.667, 1.0],
                }
            ),
            sheet_name="01_Test",
        ),
        AnalysisResult(
            name="Test Analysis 2",
            title="Test Title 2",
            df=pd.DataFrame(
                {
                    "Metric": ["Total", "Rate"],
                    "Value": [100, "50.0%"],
                }
            ),
            sheet_name="02_Test",
        ),
    ]


class TestWriteExcelReport:
    def test_creates_file(self, sample_settings, sample_df, mock_analyses, tmp_path):
        path = tmp_path / "test_report.xlsx"
        result = write_excel_report(
            sample_settings,
            sample_df,
            mock_analyses,
            output_path=path,
        )
        assert result.exists()
        assert result.suffix == ".xlsx"

    def test_auto_generates_path(self, sample_settings, sample_df, mock_analyses):
        sample_settings.output_dir.mkdir(parents=True, exist_ok=True)
        result = write_excel_report(sample_settings, sample_df, mock_analyses)
        assert result.exists()
        assert "ICS_Analysis" in result.name

    def test_contains_sheets(self, sample_settings, sample_df, mock_analyses, tmp_path):
        from openpyxl import load_workbook

        path = tmp_path / "test_report.xlsx"
        write_excel_report(sample_settings, sample_df, mock_analyses, output_path=path)

        wb = load_workbook(path)
        sheet_names = wb.sheetnames
        assert "Report Info" in sheet_names
        assert "Contents" in sheet_names
        assert "01_Test" in sheet_names
        assert "02_Test" in sheet_names
        wb.close()

    def test_skips_errored_analyses(self, sample_settings, sample_df, tmp_path):
        analyses = [
            AnalysisResult(
                name="Good",
                title="Good Analysis",
                df=pd.DataFrame({"A": [1]}),
                sheet_name="Good",
            ),
            AnalysisResult(
                name="Bad",
                title="Bad Analysis",
                df=pd.DataFrame(),
                error="Something broke",
            ),
        ]
        path = tmp_path / "test_report.xlsx"
        write_excel_report(sample_settings, sample_df, analyses, output_path=path)

        from openpyxl import load_workbook

        wb = load_workbook(path)
        assert "Good" in wb.sheetnames
        assert "Bad" not in wb.sheetnames
        wb.close()

    def test_empty_analyses(self, sample_settings, sample_df, tmp_path):
        path = tmp_path / "test_report.xlsx"
        write_excel_report(sample_settings, sample_df, [], output_path=path)
        assert path.exists()
