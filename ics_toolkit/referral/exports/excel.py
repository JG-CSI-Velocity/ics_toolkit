"""Referral intelligence Excel report -- reuses analysis export styles."""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path

import pandas as pd

from ics_toolkit.analysis.analyses.base import AnalysisResult
from ics_toolkit.analysis.exports.excel import (
    _register_styles,
    _write_analysis_sheet,
    _write_toc_sheet,
)
from ics_toolkit.settings import ReferralSettings

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
except ImportError:
    Workbook = None  # type: ignore[assignment,misc]

logger = logging.getLogger(__name__)

NAVY = "1B365D"


def _write_referral_cover(
    wb: "Workbook",
    settings: ReferralSettings,
    df: pd.DataFrame,
    analyses: list[AnalysisResult],
) -> None:
    """Write cover sheet for the referral report."""
    ws = wb.active
    ws.title = "Report Info"
    ws.sheet_properties.showGridLines = False

    ws.merge_cells("A1:D1")
    cell = ws["A1"]
    cell.value = "Referral Intelligence Report"
    cell.font = Font(name="Calibri", size=24, bold=True, color=NAVY)

    ws.merge_cells("A2:D2")
    ws["A2"].value = settings.client_name or f"Client {settings.client_id or 'unknown'}"
    ws["A2"].font = Font(name="Calibri", size=16, color="666666")

    now = datetime.now()
    details = [
        ("Client ID:", settings.client_id or "unknown"),
        ("Report Date:", now.strftime("%B %d, %Y")),
        ("Source File:", settings.data_file.name if settings.data_file else "N/A"),
        ("Total Referral Records:", f"{len(df):,}"),
        (
            "Unique Referrers:",
            (f"{df['Referrer'].nunique():,}" if "Referrer" in df.columns else "N/A"),
        ),
        ("Analyses Run:", str(sum(1 for a in analyses if a.error is None))),
    ]

    for i, (label, value) in enumerate(details, start=4):
        ws[f"A{i}"].value = label
        ws[f"A{i}"].font = Font(name="Calibri", bold=True, size=11)
        ws[f"B{i}"].value = value
        ws[f"B{i}"].font = Font(name="Calibri", size=11)

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 50


def write_referral_excel(
    settings: ReferralSettings,
    df: pd.DataFrame,
    analyses: list[AnalysisResult],
    output_path: Path | None = None,
    chart_pngs: dict[str, bytes] | None = None,
) -> Path:
    """Write the referral intelligence Excel report.

    Reuses NamedStyles and sheet writers from the analysis export module.
    """
    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = settings.output_dir / f"Referral_Intelligence_{timestamp}.xlsx"

    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    _register_styles(wb)
    _write_referral_cover(wb, settings, df, analyses)
    _write_toc_sheet(wb, analyses)

    pngs = chart_pngs or {}
    for analysis in analyses:
        if analysis.error is not None:
            continue
        _write_analysis_sheet(wb, analysis, chart_png=pngs.get(analysis.name))

    wb.save(output_path)
    logger.info("Referral Excel report saved: %s", output_path)
    return output_path
